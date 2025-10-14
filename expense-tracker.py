# expense_tracker.py - Day 10
# Added database optimization with indexes and comprehensive reporting features

from datetime import datetime, timedelta
import sqlite3
import os
import matplotlib.pyplot as plt
from pathlib import Path
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, PieChart, Reference

CATEGORIES = ["Food", "Transport", "Entertainment", "Shopping", "Bills", "Other"]
DB_FILE = "expenses.db"

def init_database():
    """Create database and expenses table if they don't exist"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            amount REAL NOT NULL,
            description TEXT NOT NULL,
            category TEXT NOT NULL,
            date TEXT NOT NULL
        )
    ''')
    
    # NEW: Create indexes for better performance
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_date ON expenses(date)
    ''')
    
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_category ON expenses(category)
    ''')
    
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_amount ON expenses(amount)
    ''')
    
    conn.commit()
    conn.close()
    print("✓ Database initialized with indexes")

def display_categories():
    """Display available categories"""
    print("\nCategories:")
    for i, category in enumerate(CATEGORIES, 1):
        print(f"{i}. {category}")

def get_category():
    """Get a valid category from user"""
    display_categories()
    
    while True:
        try:
            choice = int(input("\nSelect category (1-6): "))
            if 1 <= choice <= len(CATEGORIES):
                return CATEGORIES[choice - 1]
            else:
                print(f"Please enter a number between 1 and {len(CATEGORIES)}")
        except ValueError:
            print("Please enter a valid number")

def load_expenses():
    """Load all expenses from database"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    cursor.execute('SELECT id, amount, description, category, date FROM expenses ORDER BY date DESC')
    rows = cursor.fetchall()
    
    expenses = []
    for row in rows:
        expense = {
            "id": row[0],
            "amount": row[1],
            "description": row[2],
            "category": row[3],
            "date": datetime.strptime(row[4], "%Y-%m-%d %H:%M:%S")
        }
        expenses.append(expense)
    
    conn.close()
    
    if len(expenses) > 0:
        print(f"✓ Loaded {len(expenses)} expenses from database")
    
    return expenses

def add_expense_to_db(amount, description, category, date):
    """Add a new expense to the database"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO expenses (amount, description, category, date)
        VALUES (?, ?, ?, ?)
    ''', (amount, description, category, date.strftime("%Y-%m-%d %H:%M:%S")))
    
    conn.commit()
    expense_id = cursor.lastrowid
    conn.close()
    
    return expense_id

def delete_expense_from_db(expense_id):
    """Delete an expense from the database"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    cursor.execute('DELETE FROM expenses WHERE id = ?', (expense_id,))
    
    conn.commit()
    conn.close()

def update_expense_in_db(expense_id, amount, description, category):
    """Update an expense in the database"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    cursor.execute('''
        UPDATE expenses 
        SET amount = ?, description = ?, category = ?
        WHERE id = ?
    ''', (amount, description, category, expense_id))
    
    conn.commit()
    conn.close()

def search_expenses_in_db(search_term):
    """Search expenses by description or category"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT id, amount, description, category, date 
        FROM expenses 
        WHERE LOWER(description) LIKE LOWER(?) 
           OR LOWER(category) LIKE LOWER(?)
        ORDER BY date DESC
    ''', (f'%{search_term}%', f'%{search_term}%'))
    
    rows = cursor.fetchall()
    
    expenses = []
    for row in rows:
        expense = {
            "id": row[0],
            "amount": row[1],
            "description": row[2],
            "category": row[3],
            "date": datetime.strptime(row[4], "%Y-%m-%d %H:%M:%S")
        }
        expenses.append(expense)
    
    conn.close()
    
    return expenses

def advanced_search(min_amount=None, max_amount=None, category=None, search_term=None):
    """Advanced search with multiple filters"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    query = 'SELECT id, amount, description, category, date FROM expenses WHERE 1=1'
    params = []
    
    if min_amount is not None:
        query += ' AND amount >= ?'
        params.append(min_amount)
    
    if max_amount is not None:
        query += ' AND amount <= ?'
        params.append(max_amount)
    
    if category is not None:
        query += ' AND category = ?'
        params.append(category)
    
    if search_term is not None:
        query += ' AND LOWER(description) LIKE LOWER(?)'
        params.append(f'%{search_term}%')
    
    query += ' ORDER BY date DESC'
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    
    expenses = []
    for row in rows:
        expense = {
            "id": row[0],
            "amount": row[1],
            "description": row[2],
            "category": row[3],
            "date": datetime.strptime(row[4], "%Y-%m-%d %H:%M:%S")
        }
        expenses.append(expense)
    
    conn.close()
    
    return expenses

# NEW: Generate monthly report
def generate_monthly_report(year, month):
    """Generate a detailed report for a specific month"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Date range for the month
    start_date = f"{year}-{month:02d}-01 00:00:00"
    
    # Calculate last day of month
    if month == 12:
        end_date = f"{year + 1}-01-01 00:00:00"
    else:
        end_date = f"{year}-{month + 1:02d}-01 00:00:00"
    
    # Get all expenses for the month
    cursor.execute('''
        SELECT category, SUM(amount) as total, COUNT(*) as count, AVG(amount) as average
        FROM expenses
        WHERE date >= ? AND date < ?
        GROUP BY category
        ORDER BY total DESC
    ''', (start_date, end_date))
    
    category_stats = cursor.fetchall()
    
    # Get overall stats
    cursor.execute('''
        SELECT 
            SUM(amount) as total,
            COUNT(*) as count,
            AVG(amount) as average,
            MIN(amount) as minimum,
            MAX(amount) as maximum
        FROM expenses
        WHERE date >= ? AND date < ?
    ''', (start_date, end_date))
    
    overall_stats = cursor.fetchone()
    
    conn.close()
    
    return category_stats, overall_stats

# NEW: Generate yearly summary
def generate_yearly_summary(year):
    """Generate yearly summary with month-by-month breakdown"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    monthly_totals = []
    
    for month in range(1, 13):
        start_date = f"{year}-{month:02d}-01 00:00:00"
        
        if month == 12:
            end_date = f"{year + 1}-01-01 00:00:00"
        else:
            end_date = f"{year}-{month + 1:02d}-01 00:00:00"
        
        cursor.execute('''
            SELECT SUM(amount), COUNT(*)
            FROM expenses
            WHERE date >= ? AND date < ?
        ''', (start_date, end_date))
        
        result = cursor.fetchone()
        total = result[0] if result[0] else 0
        count = result[1] if result[1] else 0
        
        monthly_totals.append({
            'month': month,
            'total': total,
            'count': count
        })
    
    conn.close()
    
    return monthly_totals

# NEW: Get spending insights
def get_spending_insights():
    """Get insights about spending patterns"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    insights = {}
    
    # Most expensive category
    cursor.execute('''
        SELECT category, SUM(amount) as total
        FROM expenses
        GROUP BY category
        ORDER BY total DESC
        LIMIT 1
    ''')
    result = cursor.fetchone()
    if result:
        insights['highest_category'] = {'category': result[0], 'total': result[1]}
    
    # Most frequent category
    cursor.execute('''
        SELECT category, COUNT(*) as count
        FROM expenses
        GROUP BY category
        ORDER BY count DESC
        LIMIT 1
    ''')
    result = cursor.fetchone()
    if result:
        insights['frequent_category'] = {'category': result[0], 'count': result[1]}
    
    # Largest single expense
    cursor.execute('''
        SELECT amount, description, category, date
        FROM expenses
        ORDER BY amount DESC
        LIMIT 1
    ''')
    result = cursor.fetchone()
    if result:
        insights['largest_expense'] = {
            'amount': result[0],
            'description': result[1],
            'category': result[2],
            'date': result[3]
        }
    
    # Average expense
    cursor.execute('SELECT AVG(amount) FROM expenses')
    result = cursor.fetchone()
    if result and result[0]:
        insights['average_expense'] = result[0]
    
    # Total number of expenses
    cursor.execute('SELECT COUNT(*) FROM expenses')
    result = cursor.fetchone()
    if result:
        insights['total_count'] = result[0]
    
    conn.close()
    
    return insights

# NEW: Monthly report menu
def show_monthly_report():
    """Display monthly report menu"""
    try:
        year = int(input("\nEnter year (e.g., 2025): "))
        month = int(input("Enter month (1-12): "))
        
        if month < 1 or month > 12:
            print("Invalid month. Please enter 1-12")
            return
        
        category_stats, overall_stats = generate_monthly_report(year, month)
        
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December']
        
        print("\n" + "="*50)
        print(f"=== Monthly Report: {month_names[month]} {year} ===")
        print("="*50)
        
        if overall_stats[0] is None:
            print(f"\nNo expenses recorded for {month_names[month]} {year}")
            return
        
        # Overall stats
        print(f"\n📊 Overall Statistics:")
        print(f"   Total Spent: ${overall_stats[0]:.2f}")
        print(f"   Number of Expenses: {overall_stats[1]}")
        print(f"   Average Expense: ${overall_stats[2]:.2f}")
        print(f"   Smallest Expense: ${overall_stats[3]:.2f}")
        print(f"   Largest Expense: ${overall_stats[4]:.2f}")
        
        # Category breakdown
        print(f"\n📈 Category Breakdown:")
        for cat_stat in category_stats:
            category, total, count, average = cat_stat
            percentage = (total / overall_stats[0]) * 100
            print(f"   {category}:")
            print(f"      Total: ${total:.2f} ({percentage:.1f}%)")
            print(f"      Count: {count} expenses")
            print(f"      Average: ${average:.2f}")
        
    except ValueError:
        print("Invalid input. Please enter valid numbers.")

# NEW: Yearly summary menu
def show_yearly_summary():
    """Display yearly summary"""
    try:
        year = int(input("\nEnter year (e.g., 2025): "))
        
        monthly_totals = generate_yearly_summary(year)
        
        print("\n" + "="*50)
        print(f"=== Yearly Summary: {year} ===")
        print("="*50)
        
        month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        
        yearly_total = sum(m['total'] for m in monthly_totals)
        yearly_count = sum(m['count'] for m in monthly_totals)
        
        if yearly_total == 0:
            print(f"\nNo expenses recorded for {year}")
            return
        
        print(f"\n📅 Month-by-Month Breakdown:")
        print(f"{'Month':<12} {'Total':>12} {'Count':>8} {'% of Year':>12}")
        print("-" * 50)
        
        for i, month_data in enumerate(monthly_totals):
            if month_data['total'] > 0:
                percentage = (month_data['total'] / yearly_total) * 100
                print(f"{month_names[i]:<12} ${month_data['total']:>10.2f} {month_data['count']:>8} {percentage:>11.1f}%")
        
        print("-" * 50)
        print(f"{'TOTAL':<12} ${yearly_total:>10.2f} {yearly_count:>8} {'100.0%':>12}")
        print(f"\nAverage per month: ${yearly_total / 12:.2f}")
        
    except ValueError:
        print("Invalid input. Please enter a valid year.")

# NEW: Show insights menu
def show_insights():
    """Display spending insights"""
    insights = get_spending_insights()
    
    if not insights or insights.get('total_count', 0) == 0:
        print("\nNo expenses recorded yet. Add some expenses to see insights!")
        return
    
    print("\n" + "="*50)
    print("=== 💡 Spending Insights ===")
    print("="*50)
    
    print(f"\n📊 General Statistics:")
    print(f"   Total Expenses Recorded: {insights['total_count']}")
    print(f"   Average Expense: ${insights['average_expense']:.2f}")
    
    if 'highest_category' in insights:
        print(f"\n💰 Highest Spending Category:")
        print(f"   {insights['highest_category']['category']}: ${insights['highest_category']['total']:.2f}")
    
    if 'frequent_category' in insights:
        print(f"\n🔄 Most Frequent Category:")
        print(f"   {insights['frequent_category']['category']}: {insights['frequent_category']['count']} expenses")
    
    if 'largest_expense' in insights:
        print(f"\n🎯 Largest Single Expense:")
        print(f"   ${insights['largest_expense']['amount']:.2f} - {insights['largest_expense']['description']}")
        print(f"   Category: {insights['largest_expense']['category']}")
        print(f"   Date: {insights['largest_expense']['date']}")

def delete_expense(expenses):
    """Delete an expense from the list"""
    if len(expenses) == 0:
        print("\nNo expenses to delete!")
        return expenses
    
    print("\n" + "="*40)
    print("=== Select Expense to Delete ===")
    for i, expense in enumerate(expenses, 1):
        date_str = expense['date'].strftime("%Y-%m-%d %H:%M")
        print(f"{i}. ${expense['amount']:.2f} - {expense['description']} [{expense['category']}] ({date_str})")
    
    try:
        choice = int(input("\nEnter expense number to delete (0 to cancel): "))
        
        if choice == 0:
            print("Delete cancelled")
            return expenses
        
        if 1 <= choice <= len(expenses):
            deleted = expenses[choice - 1]
            
            delete_expense_from_db(deleted['id'])
            
            expenses.pop(choice - 1)
            
            print(f"✓ Deleted: ${deleted['amount']:.2f} - {deleted['description']}")
        else:
            print(f"Invalid number. Please enter 1-{len(expenses)}")
    except ValueError:
        print("Please enter a valid number")
    
    return expenses

def edit_expense(expenses):
    """Edit an existing expense"""
    if len(expenses) == 0:
        print("\nNo expenses to edit!")
        return expenses
    
    print("\n" + "="*40)
    print("=== Select Expense to Edit ===")
    for i, expense in enumerate(expenses, 1):
        date_str = expense['date'].strftime("%Y-%m-%d %H:%M")
        print(f"{i}. ${expense['amount']:.2f} - {expense['description']} [{expense['category']}] ({date_str})")
    
    try:
        choice = int(input("\nEnter expense number to edit (0 to cancel): "))
        
        if choice == 0:
            print("Edit cancelled")
            return expenses
        
        if 1 <= choice <= len(expenses):
            expense = expenses[choice - 1]
            
            print("\n" + "="*40)
            print("=== Current Expense Details ===")
            print(f"Amount: ${expense['amount']:.2f}")
            print(f"Description: {expense['description']}")
            print(f"Category: {expense['category']}")
            print(f"Date: {expense['date'].strftime('%Y-%m-%d %H:%M')}")
            
            print("\nWhat would you like to edit?")
            print("1. Amount")
            print("2. Description")
            print("3. Category")
            print("4. All of the above")
            print("0. Cancel")
            
            edit_choice = input("\nEnter your choice (0-4): ")
            
            if edit_choice == "0":
                print("Edit cancelled")
                return expenses
            
            new_amount = expense['amount']
            new_description = expense['description']
            new_category = expense['category']
            
            if edit_choice in ["1", "4"]:
                try:
                    amount_input = input(f"Enter new amount (current: ${expense['amount']:.2f}): $")
                    if amount_input.strip():
                        new_amount = float(amount_input)
                except ValueError:
                    print("Invalid amount, keeping original")
            
            if edit_choice in ["2", "4"]:
                desc_input = input(f"Enter new description (current: {expense['description']}): ")
                if desc_input.strip():
                    new_description = desc_input
            
            if edit_choice in ["3", "4"]:
                print(f"\nCurrent category: {expense['category']}")
                new_category = get_category()
            
            update_expense_in_db(expense['id'], new_amount, new_description, new_category)
            
            expense['amount'] = new_amount
            expense['description'] = new_description
            expense['category'] = new_category
            
            print(f"✓ Expense updated successfully!")
            date_str = expense['date'].strftime("%Y-%m-%d %H:%M")
            print(f"   ${expense['amount']:.2f} - {expense['description']} [{expense['category']}] ({date_str})")
        else:
            print(f"Invalid number. Please enter 1-{len(expenses)}")
    except ValueError:
        print("Please enter a valid number")
    
    return expenses

def search_expenses():
    """Search expenses by keyword"""
    search_term = input("\nEnter search term (description or category): ").strip()
    
    if not search_term:
        print("Search cancelled")
        return
    
    results = search_expenses_in_db(search_term)
    
    if len(results) == 0:
        print(f"\nNo expenses found matching '{search_term}'")
    else:
        print("\n" + "="*40)
        print(f"=== Search Results for '{search_term}' ===")
        print(f"Found {len(results)} expense(s)")
        print()
        
        total = 0
        for i, expense in enumerate(results, 1):
            date_str = expense['date'].strftime("%Y-%m-%d %H:%M")
            print(f"{i}. ${expense['amount']:.2f} - {expense['description']} [{expense['category']}] ({date_str})")
            total += expense['amount']
        
        print(f"\nTotal: ${total:.2f}")

def advanced_search_menu():
    """Advanced search with multiple filters"""
    print("\n" + "="*40)
    print("=== Advanced Search ===")
    
    min_amount = None
    max_amount = None
    category = None
    search_term = None
    
    amount_filter = input("\nFilter by amount range? (y/n): ").lower()
    if amount_filter == 'y':
        try:
            min_input = input("Minimum amount (press Enter to skip): $")
            if min_input.strip():
                min_amount = float(min_input)
            
            max_input = input("Maximum amount (press Enter to skip): $")
            if max_input.strip():
                max_amount = float(max_input)
        except ValueError:
            print("Invalid amount, skipping amount filter")
    
    cat_filter = input("Filter by category? (y/n): ").lower()
    if cat_filter == 'y':
        category = get_category()
    
    keyword_filter = input("Search in description? (y/n): ").lower()
    if keyword_filter == 'y':
        search_term = input("Enter keyword: ").strip()
    
    results = advanced_search(min_amount, max_amount, category, search_term)
    
    if len(results) == 0:
        print("\nNo expenses found matching your criteria")
    else:
        print("\n" + "="*40)
        print("=== Search Results ===")
        print(f"Found {len(results)} expense(s)")
        print()
        
        total = 0
        for i, expense in enumerate(results, 1):
            date_str = expense['date'].strftime("%Y-%m-%d %H:%M")
            print(f"{i}. ${expense['amount']:.2f} - {expense['description']} [{expense['category']}] ({date_str})")
            total += expense['amount']
        
        print(f"\nTotal: ${total:.2f}")

def view_expenses_by_date(expenses):
    """View expenses filtered by date range"""
    if len(expenses) == 0:
        print("\nNo expenses yet!")
        return
    
    print("\n" + "="*40)
    print("=== Date Range Filter ===")
    print("1. Last 7 days")
    print("2. Last 30 days")
    print("3. This month")
    print("4. Custom date range")
    print("0. Cancel")
    
    choice = input("\nEnter your choice (0-4): ")
    
    now = datetime.now()
    filtered = []
    title = ""
    
    if choice == "1":
        start_date = now - timedelta(days=7)
        filtered = [e for e in expenses if e['date'] >= start_date]
        title = "Last 7 Days"
        
    elif choice == "2":
        start_date = now - timedelta(days=30)
        filtered = [e for e in expenses if e['date'] >= start_date]
        title = "Last 30 Days"
        
    elif choice == "3":
        start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        filtered = [e for e in expenses if e['date'] >= start_date]
        title = "This Month"
        
    elif choice == "4":
        try:
            print("\nEnter start date (YYYY-MM-DD):")
            start_input = input("Start date: ")
            start_date = datetime.strptime(start_input, "%Y-%m-%d")
            
            print("Enter end date (YYYY-MM-DD):")
            end_input = input("End date: ")
            end_date = datetime.strptime(end_input, "%Y-%m-%d")
            end_date = end_date.replace(hour=23, minute=59, second=59)
            
            if start_date > end_date:
                print("Error: Start date must be before end date!")
                return
            
            filtered = [e for e in expenses if start_date <= e['date'] <= end_date]
            title = f"From {start_input} to {end_input}"
        except ValueError:
            print("Invalid date format. Please use YYYY-MM-DD (e.g., 2025-10-01)")
            return
    
    elif choice == "0":
        print("Cancelled")
        return
    else:
        print("Invalid choice")
        return
    
    if len(filtered) == 0:
        print(f"\nNo expenses found for: {title}")
    else:
        print("\n" + "="*40)
        print(f"=== Expenses: {title} ===")
        total = 0
        for i, expense in enumerate(filtered, 1):
            date_str = expense['date'].strftime("%Y-%m-%d %H:%M")
            print(f"{i}. ${expense['amount']:.2f} - {expense['description']} [{expense['category']}] ({date_str})")
            total += expense['amount']
        
        print(f"\nTotal for {title}: ${total:.2f}")
        
        print("\nCategory Breakdown:")
        for category in CATEGORIES:
            cat_expenses = [e for e in filtered if e['category'] == category]
            if len(cat_expenses) > 0:
                cat_total = sum(e['amount'] for e in cat_expenses)
                percentage = (cat_total / total) * 100
                print(f"  {category}: ${cat_total:.2f} ({percentage:.1f}%)")
def visualize_category_spending():
    """Generate a bar chart of spending by category"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Get spending by category
    cursor.execute('''
        SELECT category, SUM(amount) as total
        FROM expenses
        GROUP BY category
        HAVING total > 0
        ORDER BY total DESC
    ''')
    
    results = cursor.fetchall()
    conn.close()
    
    if not results:
        print("\n❌ No expenses found to visualize.")
        return
    
    # Prepare data for plotting
    categories = [row[0] for row in results]
    amounts = [row[1] for row in results]
    
    # Create the bar chart
    plt.figure(figsize=(10, 6))
    bars = plt.bar(categories, amounts, color='steelblue', edgecolor='navy', linewidth=1.5)
    
    # Customize the chart
    plt.xlabel('Category', fontsize=12, fontweight='bold')
    plt.ylabel('Total Amount ($)', fontsize=12, fontweight='bold')
    plt.title('Spending by Category', fontsize=14, fontweight='bold', pad=20)
    plt.xticks(rotation=45, ha='right')
    plt.grid(axis='y', alpha=0.3, linestyle='--')
    
    # Add value labels on top of bars
    for bar in bars:
        height = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2., height,
                f'${height:.2f}',
                ha='center', va='bottom', fontsize=10, fontweight='bold')
    
    plt.tight_layout()
    
    # Create charts directory if it doesn't exist
    charts_dir = Path('charts')
    charts_dir.mkdir(exist_ok=True)
    
    # Save the chart
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'charts/category_spending_{timestamp}.png'
    plt.savefig(filename, dpi=300, bbox_inches='tight')
    print(f"\n✅ Chart saved as: {filename}")
    
    # Display the chart
    plt.show()
    print("\n📊 Close the chart window to continue...")
def visualize_spending_trends():
    """Generate line chart showing spending trends over time"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    print("\n📊 Spending Trends Visualization")
    print("1. Daily spending (last 30 days)")
    print("2. Weekly spending (last 12 weeks)")
    print("3. Monthly spending (last 12 months)")
    print("0. Cancel")
    
    choice = input("\nSelect view (0-3): ").strip()
    
    if choice == "0":
        conn.close()
        return
    
    if choice == "1":
        # Daily spending for last 30 days
        cursor.execute('''
            SELECT DATE(date) as day, SUM(amount) as total
            FROM expenses
            WHERE date >= date('now', '-30 days')
            GROUP BY DATE(date)
            ORDER BY day
        ''')
        results = cursor.fetchall()
        title = "Daily Spending (Last 30 Days)"
        xlabel = "Date"
        
    elif choice == "2":
        # Weekly spending for last 12 weeks
        cursor.execute('''
            SELECT 
                strftime('%Y-W%W', date) as week,
                SUM(amount) as total
            FROM expenses
            WHERE date >= date('now', '-84 days')
            GROUP BY strftime('%Y-W%W', date)
            ORDER BY week
        ''')
        results = cursor.fetchall()
        title = "Weekly Spending (Last 12 Weeks)"
        xlabel = "Week"
        
    elif choice == "3":
        # Monthly spending for last 12 months
        cursor.execute('''
            SELECT 
                strftime('%Y-%m', date) as month,
                SUM(amount) as total
            FROM expenses
            WHERE date >= date('now', '-365 days')
            GROUP BY strftime('%Y-%m', date)
            ORDER BY month
        ''')
        results = cursor.fetchall()
        title = "Monthly Spending (Last 12 Months)"
        xlabel = "Month"
        
    else:
        print("Invalid choice")
        conn.close()
        return
    
    conn.close()
    
    if not results:
        print(f"\n❌ No data available for {title}")
        return
    
    # Prepare data
    labels = [row[0] for row in results]
    amounts = [row[1] for row in results]
    
    # Create line chart
    plt.figure(figsize=(12, 6))
    plt.plot(labels, amounts, marker='o', linewidth=2, markersize=8, 
             color='steelblue', markerfacecolor='orange', markeredgecolor='navy')
    
    # Customize chart
    plt.xlabel(xlabel, fontsize=12, fontweight='bold')
    plt.ylabel('Amount ($)', fontsize=12, fontweight='bold')
    plt.title(title, fontsize=14, fontweight='bold', pad=20)
    plt.xticks(rotation=45, ha='right')
    plt.grid(True, alpha=0.3, linestyle='--')
    
    # Add value labels on points
    for i, (label, amount) in enumerate(zip(labels, amounts)):
        plt.text(i, amount, f'${amount:.0f}', 
                ha='center', va='bottom', fontsize=9, fontweight='bold')
    
    plt.tight_layout()
    
    # Save chart
    charts_dir = Path('charts')
    charts_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'charts/spending_trends_{choice}_{timestamp}.png'
    plt.savefig(filename, dpi=300, bbox_inches='tight')
    print(f"\n✅ Chart saved as: {filename}")
    
    # Display
    plt.show()
    print("\n📊 Close the chart window to continue...")
def visualize_category_trends():
    """Show spending trends for each category over time"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Get last 6 months of data by category
    cursor.execute('''
        SELECT 
            strftime('%Y-%m', date) as month,
            category,
            SUM(amount) as total
        FROM expenses
        WHERE date >= date('now', '-180 days')
        GROUP BY strftime('%Y-%m', date), category
        ORDER BY month, category
    ''')
    
    results = cursor.fetchall()
    conn.close()
    
    if not results:
        print("\n❌ No data available for category trends")
        return
    
    # Organize data by category
    from collections import defaultdict
    category_data = defaultdict(lambda: {'months': [], 'amounts': []})
    all_months = sorted(set(row[0] for row in results))
    
    for month, category, amount in results:
        category_data[category]['months'].append(month)
        category_data[category]['amounts'].append(amount)
    
    # Create multi-line chart
    plt.figure(figsize=(12, 7))
    
    colors = ['steelblue', 'orange', 'green', 'red', 'purple', 'brown']
    
    for idx, (category, data) in enumerate(category_data.items()):
        # Create full data with zeros for missing months
        full_amounts = []
        for month in all_months:
            if month in data['months']:
                month_idx = data['months'].index(month)
                full_amounts.append(data['amounts'][month_idx])
            else:
                full_amounts.append(0)
        
        plt.plot(all_months, full_amounts, marker='o', linewidth=2, 
                label=category, color=colors[idx % len(colors)])
    
    plt.xlabel('Month', fontsize=12, fontweight='bold')
    plt.ylabel('Amount ($)', fontsize=12, fontweight='bold')
    plt.title('Spending Trends by Category (Last 6 Months)', fontsize=14, fontweight='bold', pad=20)
    plt.xticks(rotation=45, ha='right')
    plt.legend(loc='best', fontsize=10)
    plt.grid(True, alpha=0.3, linestyle='--')
    plt.tight_layout()
    
    # Save chart
    charts_dir = Path('charts')
    charts_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'charts/category_trends_{timestamp}.png'
    plt.savefig(filename, dpi=300, bbox_inches='tight')
    print(f"\n✅ Chart saved as: {filename}")
    
    plt.show()
    print("\n📊 Close the chart window to continue...")
def generate_test_data():
    """Generate sample expenses for testing visualizations"""
    print("\n⚠️  WARNING: This will add test data to your database!")
    print("This creates 60+ sample expenses across 6 months and all categories.")
    
    confirm = input("\nDo you want to continue? (yes/no): ").strip().lower()
    
    if confirm != 'yes':
        print("Cancelled.")
        return
    
    import random
    
    # Sample data for each category
    test_data = {
        "Food": [
            ("Grocery shopping", 50, 120),
            ("Restaurant dinner", 30, 80),
            ("Fast food lunch", 10, 25),
            ("Coffee shop", 5, 15),
            ("Bakery items", 8, 20)
        ],
        "Transport": [
            ("Gas/Petrol", 40, 80),
            ("Uber/Taxi", 15, 40),
            ("Bus pass", 20, 50),
            ("Parking fee", 5, 15),
            ("Auto rickshaw", 10, 30)
        ],
        "Entertainment": [
            ("Movie tickets", 15, 40),
            ("Concert/Event", 50, 150),
            ("Streaming subscription", 10, 20),
            ("Gaming", 20, 60),
            ("Sports event", 30, 100)
        ],
        "Shopping": [
            ("Clothing", 30, 150),
            ("Electronics", 100, 500),
            ("Books", 10, 40),
            ("Home items", 20, 100),
            ("Gifts", 25, 80)
        ],
        "Bills": [
            ("Electricity", 50, 120),
            ("Internet", 30, 80),
            ("Phone bill", 20, 60),
            ("Water bill", 15, 40),
            ("Rent payment", 500, 1000)
        ],
        "Other": [
            ("Medical expenses", 20, 100),
            ("Personal care", 15, 50),
            ("Gym membership", 30, 80),
            ("Insurance", 50, 150),
            ("Miscellaneous", 10, 50)
        ]
    }
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Get current date
    today = datetime.now()
    
    # Generate expenses for last 6 months
    expenses_added = 0
    
    for month_offset in range(6):
        # Calculate date for this month (going backwards)
        target_month = today - timedelta(days=30 * month_offset)
        
        # Generate 10-15 random expenses per month
        num_expenses = random.randint(10, 15)
        
        for _ in range(num_expenses):
            # Pick random category
            category = random.choice(CATEGORIES)
            
            # Pick random expense type from that category
            expense_types = test_data[category]
            description, min_amount, max_amount = random.choice(expense_types)
            
            # Random amount within range
            amount = round(random.uniform(min_amount, max_amount), 2)
            
            # Random day in that month
            day = random.randint(1, 28)
            hour = random.randint(8, 22)
            minute = random.randint(0, 59)
            
            # Create date
            expense_date = target_month.replace(day=day, hour=hour, minute=minute, second=0)
            
            # Insert into database
            cursor.execute('''
                INSERT INTO expenses (amount, description, category, date)
                VALUES (?, ?, ?, ?)
            ''', (amount, description, category, expense_date.strftime("%Y-%m-%d %H:%M:%S")))
            
            expenses_added += 1
    
    conn.commit()
    conn.close()
    
    print(f"\n✅ Successfully added {expenses_added} test expenses!")
    print("These expenses cover:")
    print("  • Last 6 months of data")
    print("  • All 6 categories")
    print("  • Realistic amounts and descriptions")
    print("\nNow try the visualization options to see the difference!")
    print("\n💡 TIP: Compare option 14 (single line) vs option 15 (multiple colored lines)")
def visualize_category_pie_chart():
    """Generate pie chart showing category distribution"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    print("\n🥧 Category Distribution Pie Chart")
    print("1. All time")
    print("2. This month")
    print("3. Last 30 days")
    print("0. Cancel")
    
    choice = input("\nSelect period (0-3): ").strip()
    
    if choice == "0":
        conn.close()
        return
    
    # Build query based on choice
    if choice == "1":
        query = '''
            SELECT category, SUM(amount) as total
            FROM expenses
            GROUP BY category
            HAVING total > 0
            ORDER BY total DESC
        '''
        title = "Spending by Category (All Time)"
        
    elif choice == "2":
        query = '''
            SELECT category, SUM(amount) as total
            FROM expenses
            WHERE strftime('%Y-%m', date) = strftime('%Y-%m', 'now')
            GROUP BY category
            HAVING total > 0
            ORDER BY total DESC
        '''
        title = "Spending by Category (This Month)"
        
    elif choice == "3":
        query = '''
            SELECT category, SUM(amount) as total
            FROM expenses
            WHERE date >= date('now', '-30 days')
            GROUP BY category
            HAVING total > 0
            ORDER BY total DESC
        '''
        title = "Spending by Category (Last 30 Days)"
        
    else:
        print("Invalid choice")
        conn.close()
        return
    
    cursor.execute(query)
    results = cursor.fetchall()
    conn.close()
    
    if not results:
        print(f"\n❌ No data available for {title}")
        return
    
    # Prepare data
    categories = [row[0] for row in results]
    amounts = [row[1] for row in results]
    
    # Create pie chart
    plt.figure(figsize=(10, 8))
    
    # Custom colors
    colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8', '#F7DC6F']
    
    # Create pie chart with percentages
    wedges, texts, autotexts = plt.pie(amounts, labels=categories, autopct='%1.1f%%',
                                         startangle=90, colors=colors,
                                         explode=[0.05] * len(categories),
                                         shadow=True)
    
    # Make percentage text bold and white
    for autotext in autotexts:
        autotext.set_color('white')
        autotext.set_fontweight('bold')
        autotext.set_fontsize(11)
    
    # Make labels bold
    for text in texts:
        text.set_fontweight('bold')
        text.set_fontsize(12)
    
    plt.title(title, fontsize=14, fontweight='bold', pad=20)
    
    # Add legend with amounts
    legend_labels = [f'{cat}: ${amt:.2f}' for cat, amt in zip(categories, amounts)]
    plt.legend(legend_labels, loc='center left', bbox_to_anchor=(1, 0, 0.5, 1))
    
    plt.tight_layout()
    
    # Save chart
    charts_dir = Path('charts')
    charts_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'charts/category_pie_{choice}_{timestamp}.png'
    plt.savefig(filename, dpi=300, bbox_inches='tight')
    print(f"\n✅ Chart saved as: {filename}")
    
    plt.show()
    print("\n📊 Close the chart window to continue...")


def visualize_stacked_bar_chart():
    """Generate stacked bar chart showing category spending over months"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Get last 6 months of data
    cursor.execute('''
        SELECT 
            strftime('%Y-%m', date) as month,
            category,
            SUM(amount) as total
        FROM expenses
        WHERE date >= date('now', '-180 days')
        GROUP BY strftime('%Y-%m', date), category
        ORDER BY month, category
    ''')
    
    results = cursor.fetchall()
    conn.close()
    
    if not results:
        print("\n❌ No data available for stacked bar chart")
        return
    
    # Organize data
    from collections import defaultdict
    months = sorted(set(row[0] for row in results))
    category_data = defaultdict(lambda: [0] * len(months))
    
    for month, category, amount in results:
        month_idx = months.index(month)
        category_data[category][month_idx] = amount
    
    # Create stacked bar chart
    plt.figure(figsize=(12, 7))
    
    colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8', '#F7DC6F']
    
    # Create bars
    bottom = [0] * len(months)
    
    for idx, category in enumerate(CATEGORIES):
        if category in category_data:
            plt.bar(months, category_data[category], bottom=bottom,
                   label=category, color=colors[idx % len(colors)])
            # Update bottom for stacking
            bottom = [b + v for b, v in zip(bottom, category_data[category])]
    
    plt.xlabel('Month', fontsize=12, fontweight='bold')
    plt.ylabel('Amount ($)', fontsize=12, fontweight='bold')
    plt.title('Monthly Spending by Category (Stacked)', fontsize=14, fontweight='bold', pad=20)
    plt.xticks(rotation=45, ha='right')
    plt.legend(loc='upper left', fontsize=10)
    plt.grid(axis='y', alpha=0.3, linestyle='--')
    plt.tight_layout()
    
    # Save chart
    charts_dir = Path('charts')
    charts_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'charts/stacked_bar_{timestamp}.png'
    plt.savefig(filename, dpi=300, bbox_inches='tight')
    print(f"\n✅ Chart saved as: {filename}")
    
    plt.show()
    print("\n📊 Close the chart window to continue...")


def visualize_comparison_chart():
    """Compare this month vs last month spending"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # This month
    cursor.execute('''
        SELECT category, SUM(amount) as total
        FROM expenses
        WHERE strftime('%Y-%m', date) = strftime('%Y-%m', 'now')
        GROUP BY category
    ''')
    this_month = dict(cursor.fetchall())
    
    # Last month
    cursor.execute('''
        SELECT category, SUM(amount) as total
        FROM expenses
        WHERE strftime('%Y-%m', date) = strftime('%Y-%m', 'now', '-1 month')
        GROUP BY category
    ''')
    last_month = dict(cursor.fetchall())
    
    conn.close()
    
    if not this_month and not last_month:
        print("\n❌ No data available for comparison")
        return
    
    # Prepare data
    categories = CATEGORIES
    this_month_values = [this_month.get(cat, 0) for cat in categories]
    last_month_values = [last_month.get(cat, 0) for cat in categories]
    
    # Create grouped bar chart
    plt.figure(figsize=(12, 7))
    
    x = range(len(categories))
    width = 0.35
    
    plt.bar([i - width/2 for i in x], last_month_values, width, 
            label='Last Month', color='lightblue', edgecolor='navy')
    plt.bar([i + width/2 for i in x], this_month_values, width,
            label='This Month', color='steelblue', edgecolor='navy')
    
    plt.xlabel('Category', fontsize=12, fontweight='bold')
    plt.ylabel('Amount ($)', fontsize=12, fontweight='bold')
    plt.title('This Month vs Last Month Comparison', fontsize=14, fontweight='bold', pad=20)
    plt.xticks(x, categories, rotation=45, ha='right')
    plt.legend()
    plt.grid(axis='y', alpha=0.3, linestyle='--')
    plt.tight_layout()
    
    # Save chart
    charts_dir = Path('charts')
    charts_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'charts/comparison_{timestamp}.png'
    plt.savefig(filename, dpi=300, bbox_inches='tight')
    print(f"\n✅ Chart saved as: {filename}")
    
    # Show summary
    this_total = sum(this_month_values)
    last_total = sum(last_month_values)
    
    print(f"\n📊 Summary:")
    print(f"   Last Month Total: ${last_total:.2f}")
    print(f"   This Month Total: ${this_total:.2f}")
    
    if last_total > 0:
        change = ((this_total - last_total) / last_total) * 100
        if change > 0:
            print(f"   Change: ⬆️ +{change:.1f}% (increased)")
        elif change < 0:
            print(f"   Change: ⬇️ {change:.1f}% (decreased)")
        else:
            print(f"   Change: ➡️ No change")
    
    plt.show()
    print("\n📊 Close the chart window to continue...")
def export_to_excel():
    """Export expenses to Excel with formatting and charts"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    print("\n📊 Export to Excel")
    print("1. All expenses")
    print("2. This month only")
    print("3. Last 30 days")
    print("4. Custom date range")
    print("0. Cancel")
    
    choice = input("\nSelect data to export (0-4): ").strip()
    
    if choice == "0":
        conn.close()
        return
    
    # Build query based on choice
    if choice == "1":
        cursor.execute('SELECT * FROM expenses ORDER BY date DESC')
        filename_suffix = "all"
        
    elif choice == "2":
        cursor.execute('''
            SELECT * FROM expenses
            WHERE strftime('%Y-%m', date) = strftime('%Y-%m', 'now')
            ORDER BY date DESC
        ''')
        filename_suffix = "this_month"
        
    elif choice == "3":
        cursor.execute('''
            SELECT * FROM expenses
            WHERE date >= date('now', '-30 days')
            ORDER BY date DESC
        ''')
        filename_suffix = "last_30_days"
        
    elif choice == "4":
        try:
            start_date = input("Enter start date (YYYY-MM-DD): ")
            end_date = input("Enter end date (YYYY-MM-DD): ")
            cursor.execute('''
                SELECT * FROM expenses
                WHERE date >= ? AND date <= ?
                ORDER BY date DESC
            ''', (start_date + ' 00:00:00', end_date + ' 23:59:59'))
            filename_suffix = f"{start_date}_to_{end_date}"
        except Exception as e:
            print(f"Error: {e}")
            conn.close()
            return
    else:
        print("Invalid choice")
        conn.close()
        return
    
    results = cursor.fetchall()
    
    if not results:
        print("\n❌ No data to export")
        conn.close()
        return
    
    # Get category summary
    cursor.execute('''
        SELECT category, SUM(amount), COUNT(*), AVG(amount)
        FROM expenses
        GROUP BY category
        ORDER BY SUM(amount) DESC
    ''')
    category_summary = cursor.fetchall()
    
    conn.close()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: All Expenses
    ws1 = wb.active
    ws1.title = "Expenses"
    
    # Headers
    headers = ['ID', 'Amount', 'Description', 'Category', 'Date']
    ws1.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    total = 0
    for row in results:
        ws1.append(row)
        total += row[1]
    
    # Add total row
    total_row = ws1.max_row + 2
    ws1[f'A{total_row}'] = 'TOTAL'
    ws1[f'A{total_row}'].font = Font(bold=True, size=12)
    ws1[f'B{total_row}'] = total
    ws1[f'B{total_row}'].font = Font(bold=True, size=12)
    ws1[f'B{total_row}'].number_format = '$#,##0.00'
    
    # Format amount column
    for row in range(2, ws1.max_row + 1):
        ws1[f'B{row}'].number_format = '$#,##0.00'
    
    # Adjust column widths
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 20
    
    # Sheet 2: Category Summary
    ws2 = wb.create_sheet(title="Summary")
    
    # Headers
    ws2.append(['Category', 'Total Amount', 'Count', 'Average'])
    
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Add summary data
    for row in category_summary:
        ws2.append(row)
    
    # Format
    for row in range(2, ws2.max_row + 1):
        ws2[f'B{row}'].number_format = '$#,##0.00'
        ws2[f'D{row}'].number_format = '$#,##0.00'
    
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 15
    
    # Add chart
    chart = BarChart()
    chart.title = "Spending by Category"
    chart.x_axis.title = "Category"
    chart.y_axis.title = "Amount ($)"
    
    data = Reference(ws2, min_col=2, min_row=1, max_row=ws2.max_row)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 20
    
    ws2.add_chart(chart, "F2")
    
    # Save file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'expense_report_{filename_suffix}_{timestamp}.xlsx'
    
    wb.save(filename)
    
    print(f"\n✅ Excel report saved: {filename}")
    print(f"   Total expenses exported: {len(results)}")
    print(f"   Total amount: ${total:.2f}")


def export_to_pdf():
    """Generate PDF report with charts"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    print("\n📄 Generate PDF Report")
    print("1. Monthly report (with charts)")
    print("2. Category summary report")
    print("3. Full expense list")
    print("0. Cancel")
    
    choice = input("\nSelect report type (0-3): ").strip()
    
    if choice == "0":
        conn.close()
        return
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if choice == "1":
        # Monthly report
        year = int(input("Enter year (e.g., 2025): "))
        month = int(input("Enter month (1-12): "))
        
        category_stats, overall_stats = generate_monthly_report(year, month)
        
        if overall_stats[0] is None:
            print("\n❌ No data for this month")
            conn.close()
            return
        
        filename = f'monthly_report_{year}_{month:02d}_{timestamp}.pdf'
        
        # Create PDF
        doc = SimpleDocTemplate(filename, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December']
        
        title = Paragraph(f"Monthly Expense Report<br/>{month_names[month]} {year}", title_style)
        story.append(title)
        story.append(Spacer(1, 0.3*inch))
        
        # Overall statistics
        data = [
            ['Metric', 'Value'],
            ['Total Spent', f'${overall_stats[0]:.2f}'],
            ['Number of Expenses', str(overall_stats[1])],
            ['Average Expense', f'${overall_stats[2]:.2f}'],
            ['Smallest Expense', f'${overall_stats[3]:.2f}'],
            ['Largest Expense', f'${overall_stats[4]:.2f}']
        ]
        
        table = Table(data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        story.append(Spacer(1, 0.5*inch))
        
        # Category breakdown
        heading = Paragraph("Category Breakdown", styles['Heading2'])
        story.append(heading)
        story.append(Spacer(1, 0.2*inch))
        
        cat_data = [['Category', 'Total', 'Count', 'Average', '% of Total']]
        for cat_stat in category_stats:
            category, total, count, average = cat_stat
            percentage = (total / overall_stats[0]) * 100
            cat_data.append([
                category,
                f'${total:.2f}',
                str(count),
                f'${average:.2f}',
                f'{percentage:.1f}%'
            ])
        
        cat_table = Table(cat_data, colWidths=[1.5*inch, 1.2*inch, 0.8*inch, 1.2*inch, 1*inch])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightgrey, colors.white])
        ]))
        
        story.append(cat_table)
        
        # Build PDF
        doc.build(story)
        
        print(f"\n✅ PDF report saved: {filename}")
        
    elif choice == "2":
        # Category summary
        cursor.execute('''
            SELECT category, SUM(amount), COUNT(*), AVG(amount)
            FROM expenses
            GROUP BY category
            ORDER BY SUM(amount) DESC
        ''')
        results = cursor.fetchall()
        
        if not results:
            print("\n❌ No data available")
            conn.close()
            return
        
        filename = f'category_summary_{timestamp}.pdf'
        
        doc = SimpleDocTemplate(filename, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        title = Paragraph("Category Summary Report", title_style)
        story.append(title)
        story.append(Spacer(1, 0.3*inch))
        
        # Data table
        data = [['Category', 'Total Amount', 'Count', 'Average']]
        total = 0
        for row in results:
            data.append([row[0], f'${row[1]:.2f}', str(row[2]), f'${row[3]:.2f}'])
            total += row[1]
        
        data.append(['TOTAL', f'${total:.2f}', '', ''])
        
        table = Table(data, colWidths=[2*inch, 1.5*inch, 1*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        doc.build(story)
        
        print(f"\n✅ PDF report saved: {filename}")
        
    elif choice == "3":
        # Full expense list
        cursor.execute('SELECT * FROM expenses ORDER BY date DESC LIMIT 100')
        results = cursor.fetchall()
        
        if not results:
            print("\n❌ No expenses to export")
            conn.close()
            return
        
        filename = f'expense_list_{timestamp}.pdf'
        
        doc = SimpleDocTemplate(filename, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        title = Paragraph("Expense List Report<br/>(Last 100 expenses)", title_style)
        story.append(title)
        story.append(Spacer(1, 0.3*inch))
        
        # Data table
        data = [['ID', 'Amount', 'Description', 'Category', 'Date']]
        total = 0
        
        for row in results[:50]:  # Limit to 50 for PDF space
            data.append([
                str(row[0]),
                f'${row[1]:.2f}',
                row[2][:30],  # Truncate long descriptions
                row[3],
                row[4][:16]  # Date without seconds
            ])
            total += row[1]
        
        table = Table(data, colWidths=[0.5*inch, 1*inch, 3*inch, 1.2*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (1, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightgrey, colors.white])
        ]))
        
        story.append(table)
        story.append(Spacer(1, 0.2*inch))
        
        note = Paragraph(f"<b>Total shown: ${total:.2f}</b> (First 50 expenses)", styles['Normal'])
        story.append(note)
        
        doc.build(story)
        
        print(f"\n✅ PDF report saved: {filename}")
    
    conn.close()
def main():
    init_database()
    
    expenses = load_expenses()
    
    print("=== Welcome to Expense Tracker ===")
    print()
    
    while True:
        print("\n" + "="*40)
        print("What would you like to do?")
        print("1. Add an expense")
        print("2. View all expenses")
        print("3. View expenses by category")
        print("4. View expenses by date range")
        print("5. Search expenses")
        print("6. Advanced search")
        print("7. Monthly report")               # NEW option
        print("8. Yearly summary")                # NEW option
        print("9. Spending insights")             # NEW option
        print("10. View total")
        print("11. Delete an expense")
        print("12. Edit an expense")
        print("13. 📊 Visualize spending by category")
        print("14. 📈 Visualize spending trends")
        print("15. 📉 Visualize category trends")
        print("16. 🥧 Pie chart - category distribution")
        print("17. 📊 Stacked bar chart - monthly breakdown")
        print("18. 🔄 Compare this month vs last month")
        print("19. 📑 Export to Excel")
        print("20. 📄 Export to PDF")
        print("21. 🧪 Generate test data (for demos)")
        print("22. Exit")
        choice = input("\nEnter your choice (1-22): ")
        
        if choice == "1":
            # Add expense
            amount = float(input("Enter amount: $"))
            description = input("Enter description: ")
            category = get_category()
            
            current_date = datetime.now()
            
            expense_id = add_expense_to_db(amount, description, category, current_date)
            
            expense = {
                "id": expense_id,
                "amount": amount,
                "description": description,
                "category": category,
                "date": current_date
            }
            expenses.insert(0, expense)
            
            formatted_date = current_date.strftime("%Y-%m-%d %H:%M")
            print(f"✓ Added and saved: ${amount} - {description} [{category}] on {formatted_date}")
            
        elif choice == "2":
            # View all expenses
            if len(expenses) == 0:
                print("\nNo expenses yet!")
            else:
                print("\n" + "="*40)
                print("=== All Expenses ===")
                for i, expense in enumerate(expenses, 1):
                    date_str = expense['date'].strftime("%Y-%m-%d %H:%M")
                    print(f"{i}. ${expense['amount']:.2f} - {expense['description']} [{expense['category']}] ({date_str})")
                    
        elif choice == "3":
            # View expenses by category
            if len(expenses) == 0:
                print("\nNo expenses yet!")
            else:
                display_categories()
                try:
                    cat_choice = int(input("\nSelect category to view (1-6): "))
                    if 1 <= cat_choice <= len(CATEGORIES):
                        selected_category = CATEGORIES[cat_choice - 1]
                        
                        filtered = [e for e in expenses if e['category'] == selected_category]
                        
                        if len(filtered) == 0:
                            print(f"\nNo expenses in '{selected_category}' category yet!")
                        else:
                            print(f"\n=== {selected_category} Expenses ===")
                            category_total = 0
                            for i, expense in enumerate(filtered, 1):
                                date_str = expense['date'].strftime("%Y-%m-%d %H:%M")
                                print(f"{i}. ${expense['amount']:.2f} - {expense['description']} ({date_str})")
                                category_total += expense['amount']
                            print(f"\n{selected_category} Total: ${category_total:.2f}")
                    else:
                        print(f"Please enter a number between 1 and {len(CATEGORIES)}")
                except ValueError:
                    print("Please enter a valid number")
        
        elif choice == "4":
            # View expenses by date range
            view_expenses_by_date(expenses)
        
        elif choice == "5":
            # Simple search
            search_expenses()
        
        elif choice == "6":
            # Advanced search
            advanced_search_menu()
        
        elif choice == "7":
            # NEW: Monthly report
            show_monthly_report()
        
        elif choice == "8":
            # NEW: Yearly summary
            show_yearly_summary()
        
        elif choice == "9":
            # NEW: Spending insights
            show_insights()
                    
        elif choice == "10":
            # View total
            if len(expenses) == 0:
                print("\nNo expenses yet!")
            else:
                total = sum(expense['amount'] for expense in expenses)
                print(f"\nTotal expenses: ${total:.2f}")
                
                print("\nBreakdown by category:")
                for category in CATEGORIES:
                    cat_expenses = [e for e in expenses if e['category'] == category]
                    if len(cat_expenses) > 0:
                        cat_total = sum(e['amount'] for e in cat_expenses)
                        percentage = (cat_total / total) * 100
                        print(f"  {category}: ${cat_total:.2f} ({percentage:.1f}%)")
        
        elif choice == "11":
            # Delete expense
            expenses = delete_expense(expenses)
        
        elif choice == "12":
            # Edit expense
            expenses = edit_expense(expenses)
         
        elif choice == "13":
            # NEW: Visualize category spending
            visualize_category_spending()
        
        elif choice == "14":
            # Visualize spending trends
            visualize_spending_trends()
        
        elif choice == "15":
            # Visualize category trends
            visualize_category_trends()
        
        elif choice == "16":
            # Pie chart
            visualize_category_pie_chart()

        elif choice == "17":
            # Stacked bar chart
            visualize_stacked_bar_chart()

        elif choice == "18":
            # Comparison chart
            visualize_comparison_chart()

        elif choice == "19":
            # Export to Excel
            export_to_excel()

        elif choice == "20":
            # Export to PDF
            export_to_pdf()

        elif choice == "21":
            # Generate test data
            generate_test_data()

        elif choice == "22":
            print("\nGoodbye! 👋")
            break
        else:
            print("\nInvalid choice. Please try again.")

if __name__ == "__main__":
    main()